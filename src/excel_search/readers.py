from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path

from .models import CellEntry
from .text import display_value, normalize_text


SUPPORTED_EXTENSIONS = frozenset({".xlsx", ".xlsm", ".xls"})


class UnsupportedWorkbookError(ValueError):
    pass


def iter_searchable_cells(path: Path) -> Iterator[CellEntry]:
    suffix = path.suffix.casefold()
    if suffix in {".xlsx", ".xlsm"}:
        yield from _iter_openpyxl(path)
        return
    if suffix == ".xls":
        yield from _iter_xlrd(path)
        return
    raise UnsupportedWorkbookError(f"不支持的 Excel 格式：{path.suffix or '无扩展名'}")


def _iter_openpyxl(path: Path) -> Iterator[CellEntry]:
    from openpyxl import load_workbook

    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        for worksheet in workbook.worksheets:
            for row_number, values in enumerate(
                worksheet.iter_rows(min_col=1, max_col=3, values_only=True), start=1
            ):
                value_a, value_b, value_c = values
                content = display_value(value_c)
                if not content:
                    continue
                yield CellEntry(
                    sheet=worksheet.title,
                    row_number=row_number,
                    value_a=display_value(value_a),
                    value_b=display_value(value_b),
                    content=content,
                    normalized=normalize_text(content),
                )
    finally:
        workbook.close()


def _iter_xlrd(path: Path) -> Iterator[CellEntry]:
    import xlrd

    workbook = xlrd.open_workbook(path, on_demand=True)
    try:
        for worksheet in workbook.sheets():
            for row_index in range(worksheet.nrows):
                values = [
                    _xlrd_display_value(workbook, worksheet.cell(row_index, column_index))
                    if column_index < worksheet.ncols
                    else ""
                    for column_index in range(3)
                ]
                content = values[2]
                if not content:
                    continue
                yield CellEntry(
                    sheet=worksheet.name,
                    row_number=row_index + 1,
                    value_a=values[0],
                    value_b=values[1],
                    content=content,
                    normalized=normalize_text(content),
                )
    finally:
        workbook.release_resources()


def _xlrd_display_value(workbook: object, cell: object) -> str:
    import xlrd

    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return ""
    if cell.ctype == xlrd.XL_CELL_DATE:
        return display_value(xlrd.xldate_as_datetime(cell.value, workbook.datemode))
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return display_value(bool(cell.value))
    return display_value(cell.value)
